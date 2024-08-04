from django.shortcuts import render, redirect
from central.models import *
from .models import *
from .decorator import is_moderator
from django.db.models import Q
from django.contrib import messages
from django.db.models import Sum, Count
from django.conf import settings
import json
from openpyxl import Workbook
from openpyxl.styles import Font
from django.db.models import Sum, Count
from django.http import FileResponse, HttpResponse
from .exporter import *

# Dependency function
# Note: Variable names are file names changing them will cause template not found error

centralmodels = {
    'publication': publication,
    'consultancy': consultancy,
    'funding': funding,
    'ipr': ipr,
    'phd': phd,
    'r1': r1,
    'r2': r2,
    'r3': r3,
    'awards': awards,
    'd1': d1
}

EXPORTERS = {
    'publication': publicationexport(),
    'consultancy': consultancyexport(),
    'funding': fundingexport(),
    'ipr': iprexport(),
    'phd': phdexport(),
    'r1': r1export(),
    'r2': r2export(),
    'r3': r3export(),
    'awards': awardsexport(),
    'd1': d1export()
}

def pendingproposals(request):
    objects = achievements.objects.filter(Q(approvalstatus='Not Approved') | Q(approvalstatus='HoD Approved'), staff=request.user)
    return render(request, 'controller/pendingproposals.html', {'objects': objects})

@is_moderator
def approvalapplication(request, category,id):
    obj = centralmodels[category].objects.get(id=id)
    return render(request, f'controller/applications/{category}_application.html', {'object': obj})

@is_moderator
def pendingapprovals(request):
    if request.user.groups.filter(name='HoD').exists():
        objects = achievements.objects.filter(Q(approvalstatus='Not Approved'), Q(department=request.user.dept)).distinct()
    else:
        objects = achievements.objects.filter(approvalstatus='HoD Approved')
    return render(request, 'controller/pendingapprovals.html', {'objects':objects})

@is_moderator
def approve(request, category, id):
    obj = centralmodels[category].objects.get(id=id)
    if request.user.groups.filter(name='HoD').exists():
        obj.hodapproval = True
    elif request.user.groups.filter(name='Controller').exists():
        obj.controller = True
    obj.save()
    messages.success(request, 'Proposal Approved!')
    return redirect('/pendingapprovals')

def achievementspage(request):
    # Header Data
    
    all_achievements = achievements.objects.filter(staff=request.user, date__range = [settings.ACADEMIC_YEAR_START, settings.ACADEMIC_YEAR_END])
    rp = rewardpoints.objects.filter(staff=request.user, department=request.user.dept, date__range = [settings.ACADEMIC_YEAR_START, settings.ACADEMIC_YEAR_END])
    
    
    category_count = all_achievements.values("category").order_by('category').annotate(Count('category'))
    categories = []
    scores_array = []
    for i in category_count:
        categories.append(i)
        
    scores = rp.values('rc').annotate(total_points=Sum('points'))
    for score in scores:
        scores_array.append(score)
        
    recent_achievements = rp.order_by('-date')
    
    context = {
        'ach_count': all_achievements.count(), 
        'pen_count':all_achievements.filter(Q(approvalstatus='Not Approved') | Q(approvalstatus='HoD Approved')), 
        'score': rp.filter(date__range = [settings.ACADEMIC_YEAR_START, settings.ACADEMIC_YEAR_END]).aggregate(Sum('points'))['points__sum'],
        
        'scoresplitup': json.dumps(scores_array),
        'categories': json.dumps(categories),
        
        'recentach': recent_achievements,
        'leaderboard': rewardpoints.objects.filter(department=request.user.dept).values('staff__first_name').annotate(Sum('points'))[:10]
        }
    # print(recent_achievements)
    return render(request, 'controller/achievements.html', context)

def reportpage(request):
    all_achievements = achievements.objects.filter(date__range = [settings.ACADEMIC_YEAR_START, settings.ACADEMIC_YEAR_END])
    rp = rewardpoints.objects.filter(date__range = [settings.ACADEMIC_YEAR_START, settings.ACADEMIC_YEAR_END])
    dpc = []
    acc = []
    department_count = all_achievements.values('department__name').annotate(Count('department'))
    achievement_count = all_achievements.values('category').annotate(Count('category')).order_by('category')
    
    staff_data = rp.values('staff__first_name', 'staff__dept__name').annotate(Sum('points')).order_by('-points__sum')
    # Contribution Countwise data
    department_leaderboard = rp.values('department__name').annotate(Count('id')).order_by('-id__count')
    
    [dpc.append(i) for i in department_count]
    [acc.append(i) for i in achievement_count]
    
    context = {
        "department_count": json.dumps(dpc),
        "achievement_count": json.dumps(acc),
        "staff_data": staff_data,
        "department_leaderboard": department_leaderboard
    }
    return render(request, 'controller/report.html', context)

# Generating Documents
# 1 - General Staff Report
def staffreport(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f"attachment; filename=staff_report.xlsx"
    
    wb = Workbook()
    ws = wb.create_sheet(title="Staff Report", index=1)
    
    model = rewardpoints.objects.filter(date__range=[settings.ACADEMIC_YEAR_START, settings.ACADEMIC_YEAR_END]).values('staff__first_name', 'staff__dept__name').annotate(Sum('points'),Count('achid')).order_by('-points__sum')
    
    
    # print(model)
    # Assigning heading
    headings = ['Name', 'Department','Total Contributions','Points']
    row=1
    for col_num, heading in enumerate(headings, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = heading
        cell.font = Font(bold=True)
    
    # Assign Data
    for _, staff in enumerate(model, 1):
        print(staff['staff__first_name'])
        row += 1
        row_data = [
            staff['staff__first_name'],
            staff['staff__dept__name'],
            staff['achid__count'],
            staff['points__sum']
        ]
        
        for col_num, rd in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = rd
    wb.save(response)
    return response 
    
    
def achievementreport(request, reporttype):
    queryset = centralmodels[reporttype].objects.filter(date__range=[settings.ACADEMIC_YEAR_START, settings.ACADEMIC_YEAR_END])
    
    data = EXPORTERS[reporttype].export(queryset)
    
    
    response = HttpResponse(data.xlsx, content_type='application/ms-excel')
    response['Content-Disposition'] = f"attachment; filename={reporttype}_report.xlsx"
    
    return response
